from datetime import datetime
import os
import time
from tkinter.ttk import Progressbar
# from tkinter.ttk import Progressbar
import webbrowser
import pandas as pd
import openpyxl
import random
from faker import Faker
from tkinter import *
from tkinter import messagebox
from tkinter.filedialog import askopenfile 


window=Tk()
window.geometry("400x300")
window.title('Data Generator')


### GUI components###
excelbtn = Button(
    window, 
    text ='Choose Excel File', 
    command = lambda:open_file()
    ) 
excelbtn.place(x=180, y=50)
label_excel_btn = Label(window, text = "Select File" )
label_excel_btn.place(x=1,y=50)

txtfld=Entry(window, text="This is Entry Widget")
txtfld.place(x=180, y=80)
txtfld.configure(state = 'disabled')

label_txtbox1 = Label(window, text = "No. Of Records to be Generated")
label_txtbox1.place(x=1,y=80)

txtbox2 = Entry(window)
txtbox2.place(x=180,y = 110)
txtbox2.configure(state = 'disabled')

label_txtbox2 = Label(window, text = "No. Of Grouping To be Done")
label_txtbox2.place(x=1,y=110)

label_drop = Label(window, text = "Select Column for Grouping")
label_drop.place(x=1,y=140)

clicked = StringVar()
clicked.set( "SELECT" )

# open_file function opens the excel file, read its columns and values from first column

def open_file():
    
    global file_name
    cols = [0]
    file_path = askopenfile(mode='r', filetypes=[('Excel Files', '*xlsx')])
    file_name = file_path.name
    excelbtn["text"] = file_name

    column_value = pd.read_excel(file_name, usecols=cols)
    for data in column_value:
        options = column_value[data]
    drop = OptionMenu( window , clicked , *options )
    txtfld.configure(state = 'normal')
    txtbox2.configure(state = 'normal')
    btn.configure(state = 'normal')
    drop.place(x=180,y=140)

    return file_name




### show function get the input values and generate the dummy data accordingly.Save the data to Excel file ###




def show():
   

    p = Progressbar(window,orient=HORIZONTAL,length=200,mode="determinate",takefocus=True)
    p.place(x=150, y=250)
    
   
    start_time = time.time()
    txt1_val = txtfld.get()
    
    number_of_duplicates = txtbox2.get()
    column_name = clicked.get()
    if txt1_val == '':
        messagebox.showerror("Error", "Enter No. of Records to be created")
    if column_name == 'SELECT':
        messagebox.showerror("Error", "Select Column Name for Grouping")
    if number_of_duplicates == '':
        messagebox.showerror("Error", "Enter No. of Grouping to be done")

   
    wb=openpyxl.Workbook()
    ws=wb.active
    fake_data=Faker()
    df=openpyxl.load_workbook(file_name)
    df_sheet=df['Sheet1']

    ColNames = {}
    Current = 0

    ### acess column name from the excel sheet ###
    for COL in df_sheet.iter_cols(1, df_sheet.max_column):
        

        time.sleep(1)
        ColNames[COL[0].value] = Current
        Current += 1
    
    if 'type' in ColNames :
        
       
        counter=0
        no_of_records = int(txt1_val) + 2
        for row_cells in df_sheet.iter_rows():

            if p['value'] < 100:
                p['value'] += 4
                
                p.step()            
                window.update()

            if counter!=0:
                ws.cell(row=1,column=counter).value = row_cells[0].value
            
            
            if row_cells[ColNames['type']].value == 'name':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.name()

            if row_cells[ColNames['type']].value == 'int':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = random.randint(0,10000)
            if row_cells[ColNames['type']].value == 'char':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = chr(random.randint(65,90))
            if row_cells[ColNames['type']].value == 'string':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.name()
            if row_cells[ColNames['type']].value == 'email':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.email()
            if row_cells[ColNames['type']].value == 'address':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.address()
            if row_cells[ColNames['type']].value == 'phone_number':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.phone_number()
            if row_cells[ColNames['type']].value == 'url':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.url()
            if row_cells[ColNames['type']].value == 'text':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.text()
            if row_cells[ColNames['type']].value == 'sentence':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.sentence()
            if row_cells[ColNames['type']].value == 'year':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.year()
            if row_cells[ColNames['type']].value == 'word':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.word()
            if row_cells[ColNames['type']].value == 'job':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.job()
            if row_cells[ColNames['type']].value == 'country':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.country()
            if row_cells[ColNames['type']].value == 'city':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.city()
            if row_cells[ColNames['type']].value == 'latitude':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.latitude()
            if row_cells[ColNames['type']].value == 'longitude':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.longitude()
            if row_cells[ColNames['type']].value == 'index':
                index_val = 0
                for i in range(2,no_of_records):
                    index_val +=1
                    ws.cell(row=i,column=counter).value = index_val + 100
            
            
            
            counter+=1
    
    else:
         messagebox.showerror('Error','No Column available with the name "type" in excel ')
    current_date = str(datetime. now(). strftime("%Y-%m-%d%I.%M.%S"))
    new_file_name = "DataPrepared-"+current_date+".xlsx"
    df_randData = pd.DataFrame(ws.values)
    new_header = df_randData.iloc[0] #grab the first row for the header
    df_randData = df_randData[1:] #take the data less the header row
    # update_progress(75)
    df_randData.columns = new_header #set the header row as the df header
    first_val = (df_randData[column_name].to_list())   
    s = pd.Series(first_val)
    repeated_data = s.repeat(int(number_of_duplicates)).to_list()
    range_row = len(df_randData.axes[0])
    df_randData[column_name] = repeated_data[0:range_row]
    writer = pd.ExcelWriter(new_file_name)

    df_randData.to_excel(writer, index = False)
    writer.save()
    # update_progress(100)
    p['value'] = 100
    messagebox.showinfo('Success','Records Created in %s Seconds'% round((time.time() - start_time)))
    window.destroy()
    webbrowser.open(new_file_name)


btn=Button(window, text="Generate Excel",command= show)
btn.place(x=180, y=180)
btn.configure(state = 'disabled')

window.mainloop()