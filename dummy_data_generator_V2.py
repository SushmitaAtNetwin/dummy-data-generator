from datetime import datetime, timedelta
import os
import time
from tkinter.ttk import Progressbar
# from tkinter.ttk import Progressbar
import webbrowser
import pandas as pd
import openpyxl
import random
from faker import Faker
from tkinter import *
from tkinter import messagebox
from tkinter.filedialog import askopenfile 
from faker_vehicle import VehicleProvider

from selenium import webdriver

from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
import time
from selenium.webdriver.common.action_chains import ActionChains


import random




list_of_university=[]

list_of_degree = []

list_of_banks = []


# web scrapping to get list of universities
def universitydata():
    os.chdir("D:/tasks")
    url = 'https://cwur.org/2021-22.php'
    # Chromedriver is just like a chrome. you can dowload latest by it website
    driver_path = os.path.join('chromedriver')
    s = Service( driver_path)
    driver = webdriver.Chrome(service=s) 
    driver.get(url)
    time.sleep(15)

    tbody = driver.find_elements(By.XPATH, "/html/body/div/div/div/div/div[2]/table/tbody")
    for p in tbody:
        atag = p.find_elements(By.TAG_NAME,"tr")

        for atags in atag:
            tdtag = atags.find_elements(By.TAG_NAME,"td")[1]
            list_of_university.append(tdtag.text)
    return list_of_university




def get_universities():
    os.chdir("D:/tasks/input_excel")
    universities = pd.read_excel("universities.xlsx")
    list_of_universities = universities['universities'].to_list()
    for university in list_of_universities:
        list_of_university.append(university)
    return list_of_university

def get_degree():
    os.chdir("D:/tasks/input_excel")
    degree = pd.read_excel("degree.xlsx")
    list_of_degrees = degree['degrees'].to_list()
    for degrees in list_of_degrees:
        list_of_degree.append(degrees)
    return list_of_degree



def get_bank():
    os.chdir("D:/tasks/input_excel")
    banks = pd.read_excel("banks.xlsx")
    list_of_bank = banks['banks'].to_list()
    for banks in list_of_bank:
        list_of_banks.append(banks)
    return list_of_banks


window=Tk()
window.geometry("400x300")
window.title('Data Generator')


### GUI components###
excelbtn = Button(
    window, 
    text ='Choose Excel File', 
    command = lambda:open_file()
    ) 
excelbtn.place(x=180, y=50)
label_excel_btn = Label(window, text = "Select File" )
label_excel_btn.place(x=1,y=50)

txtfld=Entry(window, text="This is Entry Widget")
txtfld.place(x=180, y=80)
txtfld.configure(state = 'disabled')

label_txtbox1 = Label(window, text = "No. Of Records to be Generated")
label_txtbox1.place(x=1,y=80)

txtbox2 = Entry(window)
txtbox2.place(x=180,y = 110)
txtbox2.configure(state = 'disabled')

label_txtbox2 = Label(window, text = "No. Of Grouping To be Done")
label_txtbox2.place(x=1,y=110)

label_drop = Label(window, text = "Select Column for Grouping")
label_drop.place(x=1,y=140)

clicked = StringVar()
clicked.set( "SELECT" )

# open_file function opens the excel file, read its columns and values from first column

def open_file():
    
    global file_name
    cols = [0]
    file_path = askopenfile(mode='r', filetypes=[('Excel Files', '*xlsx')])
    file_name = file_path.name
    excelbtn["text"] = file_name

    column_value = pd.read_excel(file_name, usecols=cols)
    for data in column_value:
        options = column_value[data]
    drop = OptionMenu( window , clicked , *options )
    txtfld.configure(state = 'normal')
    txtbox2.configure(state = 'normal')
    btn.configure(state = 'normal')
    drop.place(x=180,y=140)

    return file_name




### show function get the input values and generate the dummy data accordingly.Save the data to Excel file ###

def show():
   
  
    p = Progressbar(window,orient=HORIZONTAL,length=200,mode="determinate",takefocus=True)
    p.place(x=150, y=250)
    
   
    start_time = time.time()
    txt1_val = txtfld.get()
    
    number_of_duplicates = txtbox2.get()
    column_name = clicked.get()
    if txt1_val == '':
        messagebox.showerror("Error", "Enter No. of Records to be created")
    if column_name == 'SELECT':
        messagebox.showerror("Error", "Select Column Name for Grouping")
    if number_of_duplicates == '':
        messagebox.showerror("Error", "Enter No. of Grouping to be done")

   
    wb=openpyxl.Workbook()
    ws=wb.active
    fake_data=Faker()
    fake_data.add_provider(VehicleProvider)
    df=openpyxl.load_workbook(file_name)
    df_sheet=df['Sheet1']
    dataframe_1 = pd.read_excel(file_name)
    dataframe_2 = dataframe_1[dataframe_1['type'] == 'options']
    column_name_with_options = dataframe_2['col_name'].to_list()
    

    ColNames = {}
    Current = 0

    ### acess column name from the excel sheet ###
    for COL in df_sheet.iter_cols(1, df_sheet.max_column):
        

        time.sleep(1)
        ColNames[COL[0].value] = Current
        Current += 1
   
    
    if 'type' in ColNames :
        
       
        counter=0
        no_of_records = int(txt1_val) + 2
        for row_cells in df_sheet.iter_rows():

            if p['value'] < 100:
                p['value'] += 4
                
                p.step()            
                window.update()

            if counter!=0:
                ws.cell(row=1,column=counter).value = row_cells[0].value
            
            
            if row_cells[ColNames['type']].value == 'name':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.name()

            if row_cells[ColNames['type']].value == 'int':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = random.randint(0,10000)
            if row_cells[ColNames['type']].value == 'char':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = chr(random.randint(65,90))
            if row_cells[ColNames['type']].value == 'string':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.name()
            if row_cells[ColNames['type']].value == 'email':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.email()
            if row_cells[ColNames['type']].value == 'address':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.address()
            if row_cells[ColNames['type']].value == 'phone_number':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.phone_number()
            if row_cells[ColNames['type']].value == 'url':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.url()
            if row_cells[ColNames['type']].value == 'text':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.text()
            if row_cells[ColNames['type']].value == 'sentence':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.sentence()
            if row_cells[ColNames['type']].value == 'year':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.year()
            if row_cells[ColNames['type']].value == 'word':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.word()
            if row_cells[ColNames['type']].value == 'job':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.job()
            if row_cells[ColNames['type']].value == 'country':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.country()
            if row_cells[ColNames['type']].value == 'city':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.city()
            if row_cells[ColNames['type']].value == 'latitude':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.latitude()
            if row_cells[ColNames['type']].value == 'longitude':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.longitude()
            if row_cells[ColNames['type']].value == 'grade':

                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = random.choice(['A','B','C','D'])
            if row_cells[ColNames['type']].value == 'degree':

                get_degree()

                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = random.choice(list_of_degree)
            if row_cells[ColNames['type']].value == 'university':
                get_universities()
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = random.choice(list_of_university)
            if row_cells[ColNames['type']].value == 'percentage':
                for i in range(2,no_of_records):
                    a = random.uniform(30.00,100.00)
                    b = float(a/100)*100
                    c = round(b,2)
                    ws.cell(row=i,column=counter).value = str(c) + "%"
            if row_cells[ColNames['type']].value == 'company':

                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.company()
            if row_cells[ColNames['type']].value == 'joining':

                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.date_between_dates(date_start=datetime(2000,1,1), date_end=datetime(2008,12,31))
            if row_cells[ColNames['type']].value == 'leaving ':

                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.date_between_dates(date_start=datetime(2008,1,1), date_end=datetime(2019,12,31))


            if row_cells[ColNames['type']].value == 'bank':
                get_bank()

                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = random.choice(list_of_banks)

            if row_cells[ColNames['type']].value == 'amount':

                for i in range(2,no_of_records):
                     ws.cell(row=i,column=counter).value = random.randint(10000, 200000)
            
            if row_cells[ColNames['type']].value == 'tenuer':

                for i in range(2,no_of_records):
        
                    tenure = random.randint(12,36)
                    ws.cell(row=i,column=counter).value = tenure
    
            if row_cells[ColNames['type']].value == 'IR':
                for i in range(2,no_of_records):
                    tenure = random.uniform(2,12)
                    c = round(tenure,1)
                    ws.cell(row=i,column=counter).value =  c
        
            if row_cells[ColNames['type']].value == 'vehicle_name':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value =  fake_data.vehicle_category()
            if row_cells[ColNames['type']].value == 'vehicle_type':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value = fake_data.vehicle_model()
            if row_cells[ColNames['type']].value == 'vehicle_make':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value =  fake_data.vehicle_make()
            if row_cells[ColNames['type']].value == 'vehicle_model':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value =  fake_data.vehicle_make_model()
            if row_cells[ColNames['type']].value == 'color':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value =  fake_data.color_name()
            
            if row_cells[ColNames['type']].value == 'options':

                for k in column_name_with_options:
                    if row_cells[ColNames['col_name']].value == k:
                        valforoptions = dataframe_2.loc[dataframe_2['col_name']== k ]['array_values'].values[0]
                        my_list = str(valforoptions).split(",")
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value =  random.choice(my_list)

            if row_cells[ColNames['type']].value == 'rating':
                for i in range(2,no_of_records):
                    ws.cell(row=i,column=counter).value =  random.randint(1,5)
                   
        
            if row_cells[ColNames['type']].value == 'index':
                index_val = 0
                for i in range(2,no_of_records):
                    index_val +=1
                    ws.cell(row=i,column=counter).value = index_val + 100
            counter+=1

         
    else:
         messagebox.showerror('Error','No Column available with the name "type" in excel ')
   
    current_date = str(datetime. now(). strftime("%Y-%m-%d%I.%M.%S"))
    new_file_name = "DataPrepared-"+current_date+".xlsx"
    df_randData = pd.DataFrame(ws.values)
    new_header = df_randData.iloc[0] #grab the first row for the header
    df_randData = df_randData[1:] #take the data less the header row
    # update_progress(75)
    df_randData.columns = new_header #set the header row as the df header
    first_val = (df_randData[column_name].to_list())
    s = pd.Series(first_val)
    repeated_data = s.repeat(int(number_of_duplicates)).to_list()
    range_row = len(df_randData.axes[0])
    df_randData[column_name] = repeated_data[0:range_row]
    writer = pd.ExcelWriter(new_file_name)
    df_randData.to_excel(writer, index = False)
    writer.save()
    # update_progress(100)
    p['value'] = 100
    messagebox.showinfo('Success','Records Created in %s Seconds'% round((time.time() - start_time)))
    window.destroy()
    webbrowser.open(new_file_name)


btn=Button(window, text="Generate Excel",command= show)
btn.place(x=180, y=180)
btn.configure(state = 'disabled')
window.mainloop()